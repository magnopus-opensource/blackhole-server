# Copyright 2024 Magnopus LLC

# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at

#     https://www.apache.org/licenses/LICENSE-2.0

# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

from openpyxl import *
from blackhole.models import *
import pathlib
import shutil
from datetime import datetime


class SpreadsheetWriter:
    def __init__(self, workbook_path):
        self.columnMapping = {
            SLATE_DB_COL: 0,
            TAKE_NUMBER_DB_COL: 1,
            CORRECTED_SLATE_DB_COL: 2,
            CORRECTED_TAKE_DB_COL: 3,
            VALID_DB_COL: 4,
            FRAME_RATE_DB_COL: 5,
            TIMECODE_IN_FRAMES_DB_COL: 6,
            TIMECODE_IN_SMPTE_DB_COL: 7,
            TIMECODE_OUT_FRAMES_DB_COL: 8,
            TIMECODE_OUT_SMPTE_DB_COL: 9,
            MAP_DB_COL: 10,
            LEVEL_SEQUENCE_DB_COL: 11,
            LEVEL_SNAPSHOT_DB_COL: 12,
            USD_ARCHIVE_DB_COL: 13,
            DESCRIPTION_DB_COL: 14
        }

        self.workbookPath = pathlib.Path(workbook_path)

        if not self.workbookPath.exists():
            self.workbookPath.parent.mkdir(parents=True, exist_ok=True)
            self.workbook = Workbook()
            self.workbook.remove(self.workbook.active)
        else:
            self.workbook = load_workbook(self.workbookPath)

        self._empty = True

    def create_or_retrieve_sheet(self, sheet_name):
        try:
            sheet = self.workbook[sheet_name]  # Attempt to access a nonexistent sheet
            return sheet
        except KeyError:
            sheet = self.workbook.create_sheet(sheet_name)

            column_keys = list(self.columnMapping.keys())
            header_list = [schema_label_to_title(key) for key in column_keys]

            sheet.append(header_list)
            sheet.freeze_panes = "A2"

            self.workbook.save(self.workbookPath)
            self._empty = False
            return sheet

    def create_backup(self):
        if self._empty:
            return

        backup_dir_path = self.workbookPath.parent / "Spreadsheet_Backups"
        backup_dir_path.mkdir(parents=True, exist_ok=True)

        current_datetime = datetime.now()
        date_string = current_datetime.strftime("%Y-%m-%d_%H-%M-%S")
        backup_sheet_name = self.workbookPath.stem + "_" + date_string + ".xlsx"

        backup_workbook_path = backup_dir_path / backup_sheet_name

        shutil.copy2(self.workbookPath, backup_workbook_path)

    def add_or_update_take(self, take: Take, create_backup=True):
        if create_backup:
            self.create_backup()

        take_data = take.model_dump(by_alias=True, exclude_none=True)

        # Create new sheet for the data's date if no such sheet exists
        sheet = self.create_or_retrieve_sheet(str(take_data[DATE_DB_COL]))

        # Iterate through all non-header rows to see if there's already a row with the slate
        # and take number inside our data
        existing_row = None
        for row in sheet.iter_rows(min_row=2, max_col=len(self.columnMapping)):
            if row[0].value == take_data[SLATE_DB_COL] and row[1].value == take_data[TAKE_NUMBER_DB_COL]:
                existing_row = row

        if existing_row:  # Row exists, update the cells
            for key, value in take_data.items():
                if key in self.columnMapping:
                    existing_row[self.columnMapping[key]].value = value
        else:  # Row does not exist, make a new list of elements and append it
            new_row = [take_data.get(key, "") for key in self.columnMapping.keys()]
            sheet.append(new_row)

        self.workbook.save(self.workbookPath)
